"""Lecture du catalogue de profilés SZS.

Deux sources, dans cet ordre :

1. un fichier CSV normalisé, en unités SI, livré avec le paquet — c'est la
   source utilisée à l'exécution, elle ne demande aucune dépendance ;
2. le classeur ``Profilé SZS.xlsx`` d'origine, relu uniquement pour
   régénérer ce CSV (``python -m nommogramme.profils.chargeur``), ce qui
   demande ``openpyxl``.

Le classeur porte ses facteurs d'échelle dans une ligne d'en-tête séparée
(« x10⁶ » pour les inerties, « x10³ » pour les modules) plutôt que dans les
valeurs. La lecture les applique, puis convertit en unités SI.
"""

from __future__ import annotations

import csv
import math
import re
from collections.abc import Iterator
from pathlib import Path

from ..unites import mm, mm2, mm3, mm4
from .modele import Famille, Profil

__all__ = ["Catalogue", "catalogue", "charger_csv", "lire_xlsx", "ecrire_csv"]


_RACINE_PAQUET = Path(__file__).resolve().parent.parent
_CSV_DEFAUT = _RACINE_PAQUET / "data" / "profils_szs.csv"
_XLSX_DEFAUT = _RACINE_PAQUET.parents[1] / "Profilé SZS.xlsx"

_FEUILLE = "SZS C5 I_H"
_LIGNE_ENTETES = 5
_LIGNE_ECHELLES = 4
_PREMIERE_LIGNE_DONNEES = 6

# Colonne du classeur → attribut de Profil. Les noms indicés du classeur
# (« iy2 », « iz3 », « K = Ix ») sont normalisés au passage.
_CORRESPONDANCES = {
    "m": "masse",
    "A": "A",
    "Av": "Av",
    "Aw": "Aw",
    "Iy": "Iy",
    "Wely": "Wely",
    "Wply": "Wply",
    "iy2": "iy",
    "Iz": "Iz",
    "Welz": "Welz",
    "Wplz": "Wplz",
    "iz3": "iz",
    "K = Ix": "It",
    "h": "h",
    "b": "b",
    "tw": "tw",
    "tf": "tf",
    "r": "r",
    "Um": "Um",
}

# Attribut → convertisseur vers les unités SI.
_CONVERSIONS = {
    "A": mm2,
    "Av": mm2,
    "Aw": mm2,
    "Iy": mm4,
    "Iz": mm4,
    "It": mm4,
    "Wely": mm3,
    "Wply": mm3,
    "Welz": mm3,
    "Wplz": mm3,
    "iy": mm,
    "iz": mm,
    "h": mm,
    "b": mm,
    "tw": mm,
    "tf": mm,
    "r": mm,
}

_CHAMPS_CSV = [
    "nom", "famille", "masse", "A", "h", "b", "tw", "tf", "r",
    "Iy", "Iz", "Wely", "Wply", "Welz", "Wplz", "iy", "iz",
    "Um", "iz_tabule", "It", "Av", "Aw",
]

_RATIO_RAYON_PROFIL_CREUX = 2.0
"""Rayon extérieur d'un profil creux formé à chaud, en épaisseurs de paroi.

EN 10210 : le rayon extérieur des angles vaut 2t pour les sections
rectangulaires finies à chaud. Le catalogue SZS ne renseigne pas cette colonne
pour les RRW.
"""


def _cle(nom: str) -> str:
    """Clé de recherche insensible à la casse, aux espaces et aux séparateurs.

    « HEB 300 », « HEB300 » et « heb-300 » désignent le même profilé.
    """
    return re.sub(r"[\s\-_]+", "", nom).upper()


def _famille_depuis_nom(nom: str) -> Famille:
    correspondance = re.match(r"([A-Za-z]+)", nom)
    if correspondance is None:
        raise ValueError(f"Nom de profilé non reconnu : {nom!r}")
    try:
        return Famille(correspondance.group(1).upper())
    except ValueError as erreur:
        raise ValueError(
            f"Famille inconnue pour le profilé {nom!r}. "
            f"Familles connues : {', '.join(f.value for f in Famille)}"
        ) from erreur


class Catalogue:
    """Collection de profilés, indexée par nom normalisé."""

    def __init__(self, profils: list[Profil]) -> None:
        self._profils = list(profils)
        self._index = {_cle(p.nom): p for p in self._profils}

    def __getitem__(self, nom: str) -> Profil:
        try:
            return self._index[_cle(nom)]
        except KeyError:
            raise KeyError(
                f"Profilé {nom!r} absent du catalogue "
                f"({len(self._profils)} profilés disponibles)."
            ) from None

    def __contains__(self, nom: object) -> bool:
        return isinstance(nom, str) and _cle(nom) in self._index

    def __iter__(self) -> Iterator[Profil]:
        return iter(self._profils)

    def __len__(self) -> int:
        return len(self._profils)

    def __repr__(self) -> str:
        return f"<Catalogue : {len(self._profils)} profilés>"

    def famille(self, famille: Famille | str) -> list[Profil]:
        """Tous les profilés d'une famille, dans l'ordre du catalogue.

        Accepte aussi bien un membre de ``Famille`` qu'une chaîne, dans
        n'importe quelle casse.
        """
        cible = famille if isinstance(famille, Famille) else Famille(famille.upper())
        return [p for p in self._profils if p.famille is cible]

    @property
    def noms(self) -> list[str]:
        return [p.nom for p in self._profils]


def charger_csv(chemin: Path | str | None = None) -> Catalogue:
    """Charge le catalogue depuis le CSV normalisé (unités SI)."""
    chemin = Path(chemin) if chemin is not None else _CSV_DEFAUT
    if not chemin.exists():
        raise FileNotFoundError(
            f"Catalogue introuvable : {chemin}. "
            "Régénérez-le avec « python -m nommogramme.profils.chargeur »."
        )

    profils: list[Profil] = []
    with chemin.open(encoding="utf-8", newline="") as fichier:
        for ligne in csv.DictReader(fichier):
            valeurs: dict[str, object] = {
                "nom": ligne["nom"],
                "famille": Famille(ligne["famille"]),
            }
            for champ in _CHAMPS_CSV[2:]:
                brut = ligne[champ]
                valeurs[champ] = float(brut) if brut else None
            profils.append(Profil(**valeurs))  # type: ignore[arg-type]
    return Catalogue(profils)


def lire_xlsx(chemin: Path | str | None = None) -> Catalogue:
    """Lit le classeur SZS d'origine et convertit en unités SI.

    Nécessite ``openpyxl`` (extra ``[xlsx]``).
    """
    try:
        import openpyxl
    except ImportError as erreur:  # pragma: no cover - dépend de l'installation
        raise ImportError(
            "La lecture du classeur SZS demande openpyxl : "
            "pip install 'nommogramme[xlsx]'"
        ) from erreur

    chemin = Path(chemin) if chemin is not None else _XLSX_DEFAUT
    classeur = openpyxl.load_workbook(chemin, data_only=True)
    feuille = classeur[_FEUILLE]

    colonnes: dict[str, int] = {}
    echelles: dict[str, float] = {}
    for indice in range(1, feuille.max_column + 1):
        entete = feuille.cell(row=_LIGNE_ENTETES, column=indice).value
        if entete in _CORRESPONDANCES:
            attribut = _CORRESPONDANCES[entete]
            colonnes[attribut] = indice
            echelles[attribut] = _echelle(
                feuille.cell(row=_LIGNE_ECHELLES, column=indice).value
            )

    manquantes = set(_CORRESPONDANCES.values()) - set(colonnes)
    if manquantes:
        raise ValueError(
            f"Colonnes absentes de la feuille {_FEUILLE!r} : {sorted(manquantes)}"
        )

    profils: list[Profil] = []
    for ligne in range(_PREMIERE_LIGNE_DONNEES, feuille.max_row + 1):
        nom = feuille.cell(row=ligne, column=1).value
        if not nom:
            continue
        profils.append(_profil_depuis_ligne(feuille, ligne, colonnes, echelles, str(nom)))

    return Catalogue(profils)


def _echelle(cellule: object) -> float:
    """Interprète une cellule d'échelle (« x10⁶ », « x103 »…) en facteur."""
    if not isinstance(cellule, str):
        return 1.0
    correspondance = re.search(r"x\s*10\s*(\d+)", cellule.replace("⁶", "6").replace("³", "3"))
    return 10.0 ** int(correspondance.group(1)) if correspondance else 1.0


def _profil_depuis_ligne(
    feuille,
    ligne: int,
    colonnes: dict[str, int],
    echelles: dict[str, float],
    nom: str,
) -> Profil:
    nom = nom.strip()
    famille = _famille_depuis_nom(nom)

    brut: dict[str, float | None] = {}
    for attribut, indice in colonnes.items():
        valeur = feuille.cell(row=ligne, column=indice).value
        if valeur is None or not isinstance(valeur, (int, float)):
            brut[attribut] = None
            continue
        valeur = float(valeur) * echelles[attribut]
        conversion = _CONVERSIONS.get(attribut)
        brut[attribut] = conversion(valeur) if conversion else valeur

    if brut.get("r") is None and famille is Famille.RRW:
        epaisseur = brut.get("tw") or brut.get("tf")
        if epaisseur is not None:
            brut["r"] = _RATIO_RAYON_PROFIL_CREUX * epaisseur

    iz_tabule = _corriger_rayon_giration_profils_creux(brut, famille, nom)

    obligatoires = ["masse", "A", "h", "b", "tw", "tf", "r", "Iy", "Iz",
                    "Wely", "Wply", "Welz", "Wplz", "iy", "iz", "Um"]
    absentes = [champ for champ in obligatoires if brut.get(champ) is None]
    if absentes:
        raise ValueError(
            f"Profilé {nom!r} (ligne {ligne}) : valeurs obligatoires absentes : {absentes}"
        )

    return Profil(
        nom=nom,
        famille=famille,
        masse=brut["masse"],  # type: ignore[arg-type]
        A=brut["A"],  # type: ignore[arg-type]
        h=brut["h"],  # type: ignore[arg-type]
        b=brut["b"],  # type: ignore[arg-type]
        tw=brut["tw"],  # type: ignore[arg-type]
        tf=brut["tf"],  # type: ignore[arg-type]
        r=brut["r"],  # type: ignore[arg-type]
        Iy=brut["Iy"],  # type: ignore[arg-type]
        Iz=brut["Iz"],  # type: ignore[arg-type]
        Wely=brut["Wely"],  # type: ignore[arg-type]
        Wply=brut["Wply"],  # type: ignore[arg-type]
        Welz=brut["Welz"],  # type: ignore[arg-type]
        Wplz=brut["Wplz"],  # type: ignore[arg-type]
        iy=brut["iy"],  # type: ignore[arg-type]
        iz=brut["iz"],  # type: ignore[arg-type]
        Um=brut["Um"],  # type: ignore[arg-type]
        iz_tabule=iz_tabule,
        It=brut.get("It"),
        Av=brut.get("Av"),
        Aw=brut.get("Aw"),
    )


def _corriger_rayon_giration_profils_creux(
    brut: dict[str, float | None], famille: Famille, nom: str
) -> float | None:
    """Rétablit i_z pour les tubes RRW, et renvoie la valeur d'origine.

    La colonne ``iz3`` du classeur SZS est **figée à 15,0018 mm sur les 108
    lignes RRW** : c'est la valeur du premier tube de la série, RRW 40/40/3,
    recopiée en dur au lieu d'être calculée. Le contrôle i = √(I/A) de
    ``coherence.auditer`` détecte l'anomalie sur 106 des 108 lignes.

    La correction est sans ambiguïté : tous les tubes du catalogue sont
    carrés, donc I_z = I_y et i_z = i_y = √(I_z/A). La valeur tabulée est
    conservée dans ``Profil.iz_tabule`` pour mémoire.

    Elle s'applique à **toute la famille**, sans condition d'écart. Le premier
    tube de la série porte par construction la bonne valeur, et deux autres
    s'en approchent à moins de 1 % par coïncidence ; les épargner sur ce seul
    critère laisserait des valeurs fausses derrière un seuil arbitraire, alors
    que la cause est identifiée et vaut pour toutes les lignes.

    L'enjeu n'est pas cosmétique : i_z pilote l'élancement de flambement dans
    le plan faible. Pour un poteau RRW 400/400/10 de 6 m en S355, i_z = 15 mm
    au lieu de 159 mm donne λ̄ = 5,23 au lieu de 0,49, soit une résistance au
    flambement de 182 kN au lieu de 4165 kN — sous-estimée d'un facteur 23.

    Aucune correction n'est appliquée aux autres familles : l'incohérence du
    HHD 320.74 est de nature inverse — c'est son I_z tabulé qui paraît bas de
    8 %, son i_z concordant avec la géométrie — et la trancher demanderait les
    tables SZS d'origine. Elle est donc seulement signalée.
    """
    if famille is not Famille.RRW:
        return None

    iz_tabule = brut.get("iz")
    Iz, A = brut.get("Iz"), brut.get("A")
    if Iz is None or A is None or A <= 0.0:
        return None

    brut["iz"] = math.sqrt(Iz / A)
    return iz_tabule


def ecrire_csv(cat: Catalogue, chemin: Path | str | None = None) -> Path:
    """Écrit le catalogue au format CSV normalisé (unités SI)."""
    chemin = Path(chemin) if chemin is not None else _CSV_DEFAUT
    chemin.parent.mkdir(parents=True, exist_ok=True)

    with chemin.open("w", encoding="utf-8", newline="") as fichier:
        redacteur = csv.writer(fichier)
        redacteur.writerow(_CHAMPS_CSV)
        for profil in cat:
            redacteur.writerow(
                [profil.nom, profil.famille.value]
                + [
                    "" if (valeur := getattr(profil, champ)) is None else repr(valeur)
                    for champ in _CHAMPS_CSV[2:]
                ]
            )
    return chemin


def _regenerer() -> None:
    """Régénère le CSV depuis le classeur et rapporte les contrôles croisés."""
    from .coherence import auditer_catalogue
    from .geometrie import ecart_relatif_um

    cat = lire_xlsx()
    destination = ecrire_csv(cat)
    print(f"{len(cat)} profilés écrits dans {destination}")

    corriges = [p for p in cat if p.iz_tabule is not None]
    if corriges:
        print(f"\ni_z recalculé sur {len(corriges)} profilés creux "
              "(colonne figée dans le classeur) :")
        for profil in corriges[:3]:
            print(f"  {profil.nom:20s} {profil.iz_tabule * 1e3:6.2f} → "
                  f"{profil.iz * 1e3:6.2f} mm")
        if len(corriges) > 3:
            print(f"  … et {len(corriges) - 3} autres")

    ecarts = [
        (abs(ecart), profil.nom, ecart)
        for profil in cat
        if (ecart := ecart_relatif_um(profil)) is not None
    ]
    ecarts.sort(reverse=True)
    moyen = sum(e[0] for e in ecarts) / len(ecarts)
    print(f"\nContrôle croisé du périmètre contre Um : écart moyen {moyen:.2%}")
    print("Écarts les plus élevés :")
    for _, nom, ecart in ecarts[:5]:
        print(f"  {nom:20s} {ecart:+.2%}")

    anomalies = auditer_catalogue(cat)
    print(f"\nAudit de cohérence : {len(anomalies)} anomalie(s)")
    for anomalie in anomalies:
        print(f"  [{anomalie.gravite.value}] {anomalie}")


if __name__ == "__main__":  # pragma: no cover
    _regenerer()


catalogue: Catalogue
"""Catalogue chargé paresseusement au premier accès."""


def __getattr__(nom: str) -> object:
    if nom == "catalogue":
        valeur = charger_csv()
        globals()["catalogue"] = valeur
        return valeur
    raise AttributeError(f"module {__name__!r} n'a pas d'attribut {nom!r}")
